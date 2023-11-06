import os,sys
import subprocess
from openpyxl import load_workbook
from openpyxl.styles import colors, Border, Side, Font, Color, PatternFill,Alignment


def markCell_Func(cell,  clr='CC0000'):
    double = Side(border_style="double")
    border = Border(left=double,
                    right=double,
                    top=double,
                    bottom=double)
    cell.border = border
    # cell.font = Font(color="FF0000")
    # cell.fill = PatternFill("solid", fgColor="DDDDDD")
    cell.fill = PatternFill("solid", fgColor=clr)

def applyCell_HeadStyle(cell,  clr='5B9BD5'):
    double = Side(border_style="double")
    border = Border(left=double,
                    right=double,
                    top=double,
                    bottom=double)
    cell.border = border
    cell.font = Font(size=16,bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor=clr)


def sheetCompare(ws1, ws2, ws_diff, diff_row):
    # print('sheet compare')
    nRow1 = ws1.max_row
    nCol1 = ws1.max_column
    nRow2 = ws2.max_row
    nCol2 = ws2.max_column
    max_row = max(nRow1, nRow2)
    max_col = max(nCol1, nCol2)
    for x in range(1, max_row+1):
        for y in range(1, max_col+1):
            cell2 = ws2.cell(row=x, column=y)
            if x > nRow1 or y > nCol1:
                clr='CC0000'
            elif x > nRow2 or y > nCol2:
                clr='548235'
            else:
                clr='929292'
            cell1 = ws1.cell(row=x, column=y)
            if (cell1.value != cell2.value):
                markCell_Func(cell2, clr)
                ws_diff.cell(diff_row, 2).value = x
                ws_diff.cell(diff_row, 3).value = y
                ws_diff.cell(diff_row, 4).value = cell1.value
                ws_diff.cell(diff_row, 5).value = cell2.value
                diff_row = diff_row+1

    return diff_row


def xlsxFileCompae(filename1, filename2, bSaveNew=False):
    wb1 = load_workbook(filename1)
    sheetnames_1 = wb1.sheetnames
    len_1 = len(sheetnames_1)
    wb2 = load_workbook(filename2)
    sheetnames_2 = wb1.sheetnames
    file_name = os.path.basename(filename1)
    filebasename = os.path.splitext(file_name)[0]
    ws_diff = wb2.create_sheet(f'diffwith_')
    ws_diff['A1'] = 'SheetName'
    ws_diff['B1'] = 'Row'
    ws_diff['C1'] = 'Column'
    ws_diff['D1'] = f'In {filebasename}'
    ws_diff['E1'] = 'In this File'
    applyCell_HeadStyle(ws_diff['A1'])
    applyCell_HeadStyle(ws_diff['B1'])
    applyCell_HeadStyle(ws_diff['C1'])
    applyCell_HeadStyle(ws_diff['D1'])
    applyCell_HeadStyle(ws_diff['E1'])
    ws_diff.row_dimensions[1].height=32
    ws_diff.column_dimensions['A'].width=30
    ws_diff.column_dimensions['C'].width=30
    ws_diff.column_dimensions['D'].width=60
    ws_diff.column_dimensions['E'].width=60
    len_2 = len(sheetnames_2)
    min_len = min(len_1, len_2)
    diff_row = 2
    for i in range(0, min_len):
        if sheetnames_1[i] == sheetnames_2[i]:
            ws_diff.cell(diff_row, 1).value = sheetnames_2[i]
            diff_row = sheetCompare(wb1.worksheets[i], wb2.worksheets[i],ws_diff,diff_row)
        else:
            print('file {0} and {1} sheet {2} name is not Same.'.format(
                filename1, filename2, i))
            ws_diff.cell(diff_row, 1).value = sheetnames_2[i]
            ws_diff.cell(diff_row, 4).value = sheetnames_1[i]
            ws_diff.cell(diff_row, 5).value = sheetnames_2[i]
            diff_row = diff_row+1
    out_mark_xlsx_file = filename2
    if (bSaveNew):
        out_mark_xlsx_file = filename2.replace('.xlsx', '_diff.xlsx')
    # print(out_mark_xlsx_file)
    wb2.save(out_mark_xlsx_file)
    if sys.platform == 'win32':
        command=f'Start-Process -FilePath {out_mark_xlsx_file}'
        subprocess.run(["powershell", "-Command", command], capture_output=False, text=True)
    


if __name__ == '__main__':
    # 全路径是为方便在vscode中进行调试
    # file_name = 'D:/workspace/demopy/excel_flow/excel/ahb_cfg_20230925.xlsx'
    file_name1 = './excel_flow/excel/UART_CFG_XY2.xlsx'
    file_name2 = './excel_flow/excel/UART_final_202301010.xlsx'
    xlsxFileCompae(file_name1, file_name2)
